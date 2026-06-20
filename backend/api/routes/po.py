"""
Purchase Order (PO) routes — create, list, approve, reject, delete.
"""
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session
from sqlalchemy import func as sa_func

from db.database import get_db
from db.models import PurchaseOrder, User
from api.dependencies import get_current_active_user, get_current_admin

router = APIRouter(prefix="/api/po", tags=["Purchase Orders"])


# ─── Pydantic Schemas ─────────────────────────────────────────────────────────

class POCreate(BaseModel):
    item_name: str = Field(..., min_length=1, max_length=500)
    item_code: Optional[str] = Field(None, max_length=50)
    category: Optional[str] = Field(None, max_length=20)
    quantity: float = Field(..., gt=0)
    unit: str = Field(default="pcs", max_length=20)
    notes: Optional[str] = None


class POResponse(BaseModel):
    id: int
    po_number: str
    item_name: str
    item_code: Optional[str]
    category: Optional[str]
    quantity: float
    unit: str
    notes: Optional[str]
    status: str
    requested_by_id: int
    requested_by_username: Optional[str]
    approved_by_id: Optional[int]
    approved_by_username: Optional[str]
    created_at: Optional[str]

    class Config:
        from_attributes = True


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _generate_po_number(db: Session) -> str:
    """Generate a unique PO number like PO-2026-0001."""
    year = datetime.now().year
    prefix = f"PO-{year}-"

    # Count existing POs for this year
    count = (
        db.query(sa_func.count(PurchaseOrder.id))
        .filter(PurchaseOrder.po_number.like(f"{prefix}%"))
        .scalar()
    )
    return f"{prefix}{(count + 1):04d}"


# ─── Routes ───────────────────────────────────────────────────────────────────

@router.post("", response_model=POResponse, status_code=status.HTTP_201_CREATED)
def create_po(
    po_in: POCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Any authenticated user can raise a Purchase Order."""
    po = PurchaseOrder(
        po_number=_generate_po_number(db),
        item_name=po_in.item_name,
        item_code=po_in.item_code,
        category=po_in.category,
        quantity=po_in.quantity,
        unit=po_in.unit,
        notes=po_in.notes,
        status="pending",
        requested_by_id=current_user.id,
    )
    db.add(po)
    db.commit()
    db.refresh(po)
    return po.to_dict()


@router.get("")
def list_pos(
    status_filter: Optional[str] = Query(None, alias="status"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """
    List POs.
    - Admin sees ALL POs.
    - Regular user sees only their own POs.
    """
    query = db.query(PurchaseOrder)

    # Non-admins only see their own POs
    if current_user.role != "admin":
        query = query.filter(PurchaseOrder.requested_by_id == current_user.id)

    if status_filter and status_filter in ("pending", "approved", "rejected"):
        query = query.filter(PurchaseOrder.status == status_filter)

    total = query.count()
    pos = query.order_by(PurchaseOrder.created_at.desc()).offset(skip).limit(limit).all()

    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "items": [po.to_dict() for po in pos],
    }


@router.get("/stats")
def po_stats(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Get PO counts by status. Admin sees global counts; user sees their own."""
    query = db.query(PurchaseOrder)
    if current_user.role != "admin":
        query = query.filter(PurchaseOrder.requested_by_id == current_user.id)

    total = query.count()
    pending = query.filter(PurchaseOrder.status == "pending").count()
    approved = query.filter(PurchaseOrder.status == "approved").count()
    rejected = query.filter(PurchaseOrder.status == "rejected").count()

    return {
        "total": total,
        "pending": pending,
        "approved": approved,
        "rejected": rejected,
    }


# ─── Export Routes ────────────────────────────────────────────────────────────

@router.get("/export/csv")
def export_csv(
    status_filter: Optional[str] = Query(None, alias="status"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    import csv
    import io
    from fastapi.responses import StreamingResponse

    query = db.query(PurchaseOrder)
    if current_user.role != "admin":
        query = query.filter(PurchaseOrder.requested_by_id == current_user.id)
    if status_filter:
        query = query.filter(PurchaseOrder.status == status_filter)

    pos = query.order_by(PurchaseOrder.created_at.desc()).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ID", "PO Number", "Item Name", "Item Code", "Category", 
        "Quantity", "Unit", "Notes", "Status", "Requested By", "Approved By", "Created At"
    ])

    for po in pos:
        writer.writerow([
            po.id, po.po_number, po.item_name, po.item_code or "", po.category or "",
            po.quantity, po.unit, po.notes or "", po.status,
            po.requested_by.username if po.requested_by else "",
            po.approved_by.username if po.approved_by else "",
            po.created_at.strftime("%Y-%m-%d %H:%M") if po.created_at else ""
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=purchase_orders.csv"},
    )


@router.get("/export/excel")
def export_excel(
    status_filter: Optional[str] = Query(None, alias="status"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    query = db.query(PurchaseOrder)
    if current_user.role != "admin":
        query = query.filter(PurchaseOrder.requested_by_id == current_user.id)
    if status_filter:
        query = query.filter(PurchaseOrder.status == status_filter)

    pos = query.order_by(PurchaseOrder.created_at.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Orders"

    headers = [
        "ID", "PO Number", "Item Name", "Item Code", "Category", 
        "Quantity", "Unit", "Notes", "Status", "Requested By", "Approved By", "Created At"
    ]
    
    header_fill = PatternFill(start_color="14532D", end_color="14532D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for po in pos:
        ws.append([
            po.id, po.po_number, po.item_name, po.item_code, po.category,
            po.quantity, po.unit, po.notes, po.status,
            po.requested_by.username if po.requested_by else None,
            po.approved_by.username if po.approved_by else None,
            po.created_at.strftime("%Y-%m-%d %H:%M") if po.created_at else None
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=purchase_orders.xlsx"},
    )


@router.get("/{po_id}", response_model=POResponse)
def get_po(
    po_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_active_user),
):
    """Get a single PO. Owner or admin can access."""
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == po_id).first()
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")

    if current_user.role != "admin" and po.requested_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to view this PO")

    return po.to_dict()


@router.put("/{po_id}/approve", response_model=POResponse)
def approve_po(
    po_id: int,
    db: Session = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
):
    """Admin approves a pending PO."""
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == po_id).first()
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")

    if po.status != "pending":
        raise HTTPException(
            status_code=400,
            detail=f"Cannot approve a PO that is already '{po.status}'",
        )

    po.status = "approved"
    po.approved_by_id = current_admin.id
    db.commit()
    db.refresh(po)
    return po.to_dict()


@router.put("/{po_id}/reject", response_model=POResponse)
def reject_po(
    po_id: int,
    db: Session = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
):
    """Admin rejects a pending PO."""
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == po_id).first()
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")

    if po.status != "pending":
        raise HTTPException(
            status_code=400,
            detail=f"Cannot reject a PO that is already '{po.status}'",
        )

    po.status = "rejected"
    po.approved_by_id = current_admin.id
    db.commit()
    db.refresh(po)
    return po.to_dict()


@router.delete("/{po_id}")
def delete_po(
    po_id: int,
    db: Session = Depends(get_db),
    current_admin: User = Depends(get_current_admin),
):
    """Admin can delete a PO record."""
    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == po_id).first()
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")

    db.delete(po)
    db.commit()
    return {"message": f"PO {po.po_number} deleted"}
