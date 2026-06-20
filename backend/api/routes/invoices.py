"""
Invoice CRUD routes + search + CSV/Excel export.
"""
import csv
import io
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import or_

from db.database import get_db
from db.repository import InvoiceRepository
from api.dependencies import get_current_admin
from pydantic import BaseModel
import uuid

router = APIRouter(prefix="/api/invoices", tags=["invoices"])

class ManualInvoiceCreate(BaseModel):
    invoice_number: str
    invoice_date: str
    order_id: Optional[str] = None
    seller_gstin: Optional[str] = None
    product_description: str
    quantity: float
    grand_total: float
    category: Optional[str] = "Category 1"


@router.get("")
def list_invoices(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    status: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
):
    repo = InvoiceRepository(db)
    invoices, total = repo.list_invoices(
        skip=skip, limit=limit, status=status, search=search
    )
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "items": [
            {
                **inv.to_dict(),
                "line_items": [li.to_dict() for li in inv.line_items],
                "taxes": [t.to_dict() for t in inv.taxes],
            }
            for inv in invoices
        ],
    }

@router.get("/advanced-search")
def advanced_search_invoices(
    query: Optional[str] = None,
    invoice_category: Optional[str] = None,
    po_category: Optional[str] = None,
    item_code: Optional[str] = None,
    process_name: Optional[str] = None,
    status: Optional[str] = None,
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db)
):
    from db.models import Invoice, PurchaseOrder, LineItem, InvoiceProcessTracking, WorkflowProcess
    
    q = db.query(Invoice)\
        .outerjoin(PurchaseOrder, Invoice.po_id == PurchaseOrder.id)\
        .outerjoin(LineItem, Invoice.id == LineItem.invoice_id)\
        .outerjoin(InvoiceProcessTracking, Invoice.id == InvoiceProcessTracking.invoice_id)\
        .outerjoin(WorkflowProcess, InvoiceProcessTracking.process_id == WorkflowProcess.id)
    
    if query:
        pattern = f"%{query}%"
        q = q.filter(
            or_(
                Invoice.invoice_number.ilike(pattern),
                Invoice.order_id.ilike(pattern),
                Invoice.seller_gstin.ilike(pattern),
                PurchaseOrder.po_number.ilike(pattern),
                PurchaseOrder.item_name.ilike(pattern),
                LineItem.name.ilike(pattern)
            )
        )
        
    if invoice_category:
        q = q.filter(Invoice.category.ilike(f"%{invoice_category}%"))
    if po_category:
        q = q.filter(PurchaseOrder.category.ilike(f"%{po_category}%"))
    if item_code:
        q = q.filter(PurchaseOrder.item_code.ilike(f"%{item_code}%"))
    if process_name:
        q = q.filter(WorkflowProcess.name.ilike(f"%{process_name}%"))
    if status:
        q = q.filter(Invoice.status.ilike(f"%{status}%"))
        
    q = q.distinct()
    
    total = q.count()
    invoices = q.order_by(Invoice.created_at.desc()).offset(skip).limit(limit).all()
    
    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "items": [
            {
                **inv.to_dict(),
                "line_items": [li.to_dict() for li in inv.line_items],
                "taxes": [t.to_dict() for t in inv.taxes],
                "tracking": [trk.to_dict() for trk in db.query(InvoiceProcessTracking).filter(InvoiceProcessTracking.invoice_id == inv.id).all()]
            }
            for inv in invoices
        ]
    }

@router.get("/unmatched")
def list_unmatched_invoices(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
):
    """List all invoices that do not have a linked PO yet."""
    from db.models import Invoice
    total = db.query(Invoice).filter(Invoice.po_id == None).count()
    invoices = db.query(Invoice).filter(Invoice.po_id == None).order_by(Invoice.created_at.desc()).offset(skip).limit(limit).all()

    return {
        "total": total,
        "skip": skip,
        "limit": limit,
        "items": [inv.to_dict() for inv in invoices],
    }


@router.post("/manual")
def create_manual_invoice(
    data: ManualInvoiceCreate,
    db: Session = Depends(get_db)
):
    from db.models import Invoice, LineItem
    # Create the manual invoice
    new_invoice = Invoice(
        file_name="Manual Entry",
        file_path="manual",
        file_hash=uuid.uuid4().hex,
        source_type="manual",
        invoice_number=data.invoice_number,
        invoice_date=data.invoice_date,
        order_id=data.order_id,
        seller_gstin=data.seller_gstin,
        grand_total=data.grand_total,
        category=data.category,
        confidence_score=100.0,
        status="processed"
    )
    db.add(new_invoice)
    db.commit()
    db.refresh(new_invoice)
    
    # Add a single line item for the product description and quantity
    line_item = LineItem(
        invoice_id=new_invoice.id,
        name=data.product_description,
        quantity=data.quantity,
        total_price=data.grand_total
    )
    db.add(line_item)
    db.commit()
    db.refresh(new_invoice)
    
    return new_invoice.to_dict()


# ── Export routes MUST be declared before /{invoice_id} to avoid
#    FastAPI matching "export" as an invoice_id path parameter. ────────────────

@router.get("/export/csv")
def export_csv(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
):
    repo = InvoiceRepository(db)
    invoices = repo.get_all_for_export(status=status)

    output = io.StringIO()
    writer = csv.writer(output)

    headers = [
        "ID", "Invoice Number", "Invoice Date",
        "Order ID", "Seller GSTIN",
        "Product Description", "HSN Code",
        "CGST Rate", "CGST Amount",
        "SGST Rate", "SGST Amount",
        "IGST Rate", "IGST Amount",
        "Total Tax", "Grand Total",
        "Confidence Score", "Status",
        "Linked PO Number", "Linked PO Item", "Linked PO Qty", "Linked PO Status",
    ]
    writer.writerow(headers)

    for inv in invoices:
        # Flatten line items into a single string
        product_desc = "; ".join(
            li.name for li in inv.line_items if li.name
        ) or ""
        hsn_codes = "; ".join(
            li.hsn_code for li in inv.line_items if li.hsn_code
        ) or ""

        # Build tax lookup
        tax_map = {t.tax_type: t for t in inv.taxes}
        cgst = tax_map.get("CGST")
        sgst = tax_map.get("SGST")
        igst = tax_map.get("IGST")
        total_tax = sum(t.amount or 0 for t in inv.taxes)

        writer.writerow([
            inv.id, inv.invoice_number,
            inv.invoice_date, inv.order_id, inv.seller_gstin,
            product_desc, hsn_codes,
            cgst.rate if cgst else "", cgst.amount if cgst else "",
            sgst.rate if sgst else "", sgst.amount if sgst else "",
            igst.rate if igst else "", igst.amount if igst else "",
            total_tax if total_tax else "",
            inv.grand_total, inv.confidence_score, inv.status,
            inv.po.po_number if inv.po else "",
            inv.po.item_name if inv.po else "",
            inv.po.quantity if inv.po else "",
            inv.po.status if inv.po else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=invoices.csv"},
    )


@router.get("/export/excel")
def export_excel(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    repo = InvoiceRepository(db)
    invoices = repo.get_all_for_export(status=status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoices"

    headers = [
        "ID", "Invoice Number", "Invoice Date",
        "Order ID", "Seller GSTIN",
        "Product Description", "HSN Code",
        "CGST Rate", "CGST Amount",
        "SGST Rate", "SGST Amount",
        "IGST Rate", "IGST Amount",
        "Total Tax", "Grand Total",
        "Confidence Score", "Status",
        "Linked PO Number", "Linked PO Item", "Linked PO Qty", "Linked PO Status",
    ]

    header_fill = PatternFill(start_color="14532D", end_color="14532D", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for inv in invoices:
        # Flatten line items
        product_desc = "; ".join(
            li.name for li in inv.line_items if li.name
        ) or ""
        hsn_codes = "; ".join(
            li.hsn_code for li in inv.line_items if li.hsn_code
        ) or ""

        # Build tax lookup
        tax_map = {t.tax_type: t for t in inv.taxes}
        cgst = tax_map.get("CGST")
        sgst = tax_map.get("SGST")
        igst = tax_map.get("IGST")
        total_tax = sum(t.amount or 0 for t in inv.taxes)

        ws.append([
            inv.id, inv.invoice_number,
            inv.invoice_date, inv.order_id, inv.seller_gstin,
            product_desc, hsn_codes,
            cgst.rate if cgst else None, cgst.amount if cgst else None,
            sgst.rate if sgst else None, sgst.amount if sgst else None,
            igst.rate if igst else None, igst.amount if igst else None,
            total_tax if total_tax else None,
            inv.grand_total, inv.confidence_score, inv.status,
            inv.po.po_number if inv.po else None,
            inv.po.item_name if inv.po else None,
            inv.po.quantity if inv.po else None,
            inv.po.status if inv.po else None,
        ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=invoices.xlsx"},
    )


# ── Parameterized routes (must come after literal paths like /export/*) ──────

@router.get("/{invoice_id}")
def get_invoice(invoice_id: int, db: Session = Depends(get_db)):
    repo = InvoiceRepository(db)
    inv = repo.get_invoice(invoice_id)
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {
        **inv.to_dict(),
        "raw_text": inv.raw_text,
        "line_items": [li.to_dict() for li in inv.line_items],
        "taxes": [t.to_dict() for t in inv.taxes],
    }


@router.put("/{invoice_id}")
def update_invoice(invoice_id: int, updates: dict, db: Session = Depends(get_db)):
    repo = InvoiceRepository(db)
    inv = repo.update_invoice(invoice_id, updates)
    if not inv:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return inv.to_dict()


@router.delete("/{invoice_id}", dependencies=[Depends(get_current_admin)])
def delete_invoice(invoice_id: int, db: Session = Depends(get_db)):
    repo = InvoiceRepository(db)
    ok = repo.delete_invoice(invoice_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Invoice not found")
    return {"message": "Deleted"}


@router.get("/{invoice_id}/suggest-po")
def suggest_po_for_invoice(invoice_id: int, db: Session = Depends(get_db)):
    """Suggest best matching POs for this invoice."""
    import difflib
    from db.models import Invoice, PurchaseOrder
    
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    # Get all POs that are approved and NOT already linked to another invoice
    linked_po_ids = db.query(Invoice.po_id).filter(Invoice.po_id.isnot(None)).subquery()
    available_pos = db.query(PurchaseOrder).filter(
        PurchaseOrder.status == "approved",
        ~PurchaseOrder.id.in_(linked_po_ids)
    ).all()

    suggestions = []
    inv_desc = (invoice.to_dict().get("product_description") or "").lower()
    inv_qty = invoice.to_dict().get("quantity") or 0
    inv_order_id = (invoice.order_id or "").lower()
    inv_number = (invoice.invoice_number or "").lower()

    for po in available_pos:
        score = 0
        po_num = po.po_number.lower()
        
        # 1. Exact PO Number match (either in order_id or invoice_number)
        if po_num == inv_order_id or po_num == inv_number or po_num in inv_desc:
            score = 100
        else:
            # 2. Keyword match (handles underscores and separate words)
            po_name_clean = po.item_name.replace("_", " ").lower()
            po_words = set(w for w in po_name_clean.split() if len(w) > 1)
            inv_words = set(w for w in inv_desc.replace("-", " ").replace("_", " ").split() if len(w) > 1)
            
            if po_words:
                common_words = po_words.intersection(inv_words)
                word_match_ratio = len(common_words) / len(po_words)
                score += word_match_ratio * 80  # Max 80 points for name match
            
            # 3. Quantity match
            if inv_qty and po.quantity and abs(inv_qty - po.quantity) < 0.01:
                score += 20  # 20 points for exact quantity match
                
        if score >= 40:
            suggestions.append({
                "po": po.to_dict(),
                "score": round(score, 1)
            })

    # Sort by score descending
    suggestions.sort(key=lambda x: x["score"], reverse=True)
    
    return {"suggestions": suggestions[:5]}  # Return top 5


@router.put("/{invoice_id}/link-po")
def link_po_to_invoice(invoice_id: int, po_id: int = Query(...), db: Session = Depends(get_db)):
    """Link a specific PO to this invoice."""
    from db.models import Invoice, PurchaseOrder
    
    invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(status_code=404, detail="Invoice not found")

    po = db.query(PurchaseOrder).filter(PurchaseOrder.id == po_id).first()
    if not po:
        raise HTTPException(status_code=404, detail="Purchase Order not found")
        
    if po.status != "approved":
        raise HTTPException(status_code=400, detail="Can only link to approved POs")

    # Check if PO is already linked
    existing_link = db.query(Invoice).filter(Invoice.po_id == po_id, Invoice.id != invoice_id).first()
    if existing_link:
        raise HTTPException(status_code=400, detail=f"PO already linked to invoice #{existing_link.id}")

    invoice.po_id = po.id
    db.commit()
    db.refresh(invoice)
    
    return invoice.to_dict()



